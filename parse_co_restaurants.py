#!/usr/bin/env python3
import json
import re
import openpyxl

def slugify(text):
    if not text:
        return ""
    text = str(text).lower()
    text = re.sub(r'[^a-z0-9]+', '-', text)
    text = re.sub(r'-+', '-', text)
    text = text.strip('-')
    return text

def parse_json_field(value):
    if value is None:
        return None
    if isinstance(value, (dict, list)):
        return value
    try:
        return json.loads(str(value))
    except (json.JSONDecodeError, TypeError):
        return None

def parse_amenities(about_data):
    amenities = {
        "delivery": False,
        "takeout": False,
        "dineIn": False,
        "outdoorSeating": False,
        "alcohol": False,
        "veganOptions": False,
        "vegetarianOptions": False,
        "acceptsReservations": False,
        "wheelchairAccessible": False,
        "casual": False,
        "cozy": False,
        "trendy": False,
        "familyFriendly": False,
        "parking": False,
        "creditCards": False
    }
    if not about_data or not isinstance(about_data, dict):
        return amenities

    def get_nested(data, category, key):
        cat = data.get(category, {})
        if isinstance(cat, dict):
            return bool(cat.get(key, False))
        return False

    amenities["delivery"] = get_nested(about_data, "Service options", "Delivery")
    amenities["takeout"] = get_nested(about_data, "Service options", "Takeout")
    amenities["dineIn"] = get_nested(about_data, "Service options", "Dine-in")
    amenities["outdoorSeating"] = get_nested(about_data, "Service options", "Outdoor seating")
    amenities["alcohol"] = get_nested(about_data, "Offerings", "Alcohol")
    amenities["veganOptions"] = get_nested(about_data, "Offerings", "Vegan options")
    amenities["vegetarianOptions"] = get_nested(about_data, "Offerings", "Vegetarian options")
    amenities["acceptsReservations"] = get_nested(about_data, "Planning", "Accepts reservations")
    amenities["wheelchairAccessible"] = get_nested(about_data, "Accessibility", "Wheelchair accessible entrance")
    amenities["casual"] = get_nested(about_data, "Atmosphere", "Casual")
    amenities["cozy"] = get_nested(about_data, "Atmosphere", "Cozy")
    amenities["trendy"] = get_nested(about_data, "Atmosphere", "Trendy")
    amenities["familyFriendly"] = get_nested(about_data, "Crowd", "Family-friendly")
    amenities["parking"] = get_nested(about_data, "Parking", "Parking")
    amenities["creditCards"] = get_nested(about_data, "Payments", "Credit cards")

    return amenities

# Load existing JSON
print("Loading existing restaurants.json...")
with open('/home/user/RamenNearYou.com/lib/restaurants.json', 'r', encoding='utf-8') as f:
    restaurants = json.load(f)

print(f"Existing records: {len(restaurants)}")

# Build lookup sets for deduplication
existing_slug_state = set()
existing_place_ids = set()
slug_city_state_counts = {}  # (slug, citySlug, stateSlug) -> count for conflict resolution

for r in restaurants:
    slug = r.get('slug', '')
    state_slug = r.get('stateSlug', '')
    city_slug = r.get('citySlug', '')
    place_id = r.get('placeId', '')

    if slug and state_slug:
        existing_slug_state.add((slug, state_slug))
    if place_id:
        existing_place_ids.add(place_id)

    key = (slug, city_slug, state_slug)
    slug_city_state_counts[key] = slug_city_state_counts.get(key, 0) + 1

# Load Excel
print("Loading Excel file...")
wb = openpyxl.load_workbook(
    '/root/.claude/uploads/e9cd06e3-4f85-4c54-8b93-20beac932830/49183abc-RamenNearYou__CO.xlsx',
    read_only=True, data_only=True
)
ws = wb.active

# Get headers from row 1
headers = []
for cell in next(ws.iter_rows(min_row=1, max_row=1)):
    headers.append(cell.value)

print(f"Columns found: {headers}")

# Map column names to indices
col_idx = {h: i for i, h in enumerate(headers) if h is not None}

def get_col(row_values, col_name):
    idx = col_idx.get(col_name)
    if idx is None:
        return None
    val = row_values[idx]
    return val

added = 0
skipped = 0

# Process data rows
for row in ws.iter_rows(min_row=2, values_only=True):
    # Skip empty rows
    if all(v is None for v in row):
        continue

    name = get_col(row, 'name')
    if not name:
        skipped += 1
        continue

    name = str(name).strip()
    city = get_col(row, 'city')
    state = get_col(row, 'state')
    place_id = get_col(row, 'place_id')

    city = str(city).strip() if city else ''
    state = str(state).strip() if state else ''
    place_id = str(place_id).strip() if place_id else ''

    # Check deduplication by placeId
    if place_id and place_id in existing_place_ids:
        skipped += 1
        continue

    base_slug = slugify(name)
    city_slug = slugify(city)
    state_slug = slugify(state)

    # Check deduplication by slug + stateSlug
    if (base_slug, state_slug) in existing_slug_state:
        skipped += 1
        continue

    # Handle slug conflicts for same city+state
    slug = base_slug
    key = (slug, city_slug, state_slug)
    count = slug_city_state_counts.get(key, 0)
    if count > 0:
        # Find a unique slug
        n = 2
        while True:
            candidate = f"{base_slug}-{n}"
            candidate_key = (candidate, city_slug, state_slug)
            if slug_city_state_counts.get(candidate_key, 0) == 0:
                slug = candidate
                key = candidate_key
                break
            n += 1

    # Parse fields
    phone = get_col(row, 'phone')
    website = get_col(row, 'website')
    address = get_col(row, 'address')
    street = get_col(row, 'street')
    county = get_col(row, 'county')
    state_code = get_col(row, 'state_code')
    postal_code = get_col(row, 'postal_code')
    latitude = get_col(row, 'latitude')
    longitude = get_col(row, 'longitude')
    rating = get_col(row, 'rating')
    reviews = get_col(row, 'reviews')
    reviews_per_score_raw = get_col(row, 'reviews_per_score')
    photos_count = get_col(row, 'photos_count')
    photo = get_col(row, 'photo')
    logo = get_col(row, 'logo')
    business_status = get_col(row, 'business_status')
    working_hours_raw = get_col(row, 'working_hours')
    price_range = get_col(row, 'range')
    description = get_col(row, 'description')
    menu_link = get_col(row, 'menu_link')
    order_links = get_col(row, 'order_links')
    location_link = get_col(row, 'location_link')
    subtypes = get_col(row, 'subtypes')
    about_raw = get_col(row, 'about')

    # Parse JSON fields
    reviews_per_score = parse_json_field(reviews_per_score_raw) or {}
    hours = parse_json_field(working_hours_raw) or {}
    about_data = parse_json_field(about_raw)
    amenities = parse_amenities(about_data)

    # Convert numeric fields
    try:
        latitude = float(latitude) if latitude is not None else None
    except (ValueError, TypeError):
        latitude = None
    try:
        longitude = float(longitude) if longitude is not None else None
    except (ValueError, TypeError):
        longitude = None
    try:
        rating = float(rating) if rating is not None else None
    except (ValueError, TypeError):
        rating = None
    try:
        reviews = int(reviews) if reviews is not None else None
    except (ValueError, TypeError):
        reviews = None
    try:
        photos_count = int(photos_count) if photos_count is not None else None
    except (ValueError, TypeError):
        photos_count = None

    record = {
        "name": name,
        "slug": slug,
        "citySlug": city_slug,
        "stateSlug": state_slug,
        "phone": str(phone).strip() if phone else "",
        "website": str(website).strip() if website else "",
        "address": str(address).strip() if address else "",
        "street": str(street).strip() if street else "",
        "city": city,
        "county": str(county).strip() if county else "",
        "state": state,
        "stateCode": str(state_code).strip() if state_code else "",
        "postalCode": str(postal_code).strip() if postal_code else "",
        "latitude": latitude,
        "longitude": longitude,
        "rating": rating,
        "reviewCount": reviews,
        "reviewsPerScore": reviews_per_score,
        "photosCount": photos_count,
        "photo": str(photo).strip() if photo else "",
        "logo": str(logo).strip() if logo else "",
        "businessStatus": str(business_status).strip() if business_status else "",
        "hours": hours,
        "priceRange": str(price_range).strip() if price_range else "",
        "description": str(description).strip() if description else "",
        "menuLink": str(menu_link).strip() if menu_link else "",
        "orderLinks": str(order_links).strip() if order_links else "",
        "googleMapsLink": str(location_link).strip() if location_link else "",
        "placeId": place_id,
        "subtypes": str(subtypes).strip() if subtypes else "",
        "amenities": amenities
    }

    restaurants.append(record)

    # Update tracking sets
    existing_slug_state.add((slug, state_slug))
    if place_id:
        existing_place_ids.add(place_id)
    slug_city_state_counts[key] = slug_city_state_counts.get(key, 0) + 1

    added += 1

wb.close()

print(f"Writing updated JSON ({len(restaurants)} total records)...")
with open('/home/user/RamenNearYou.com/lib/restaurants.json', 'w', encoding='utf-8') as f:
    json.dump(restaurants, f, ensure_ascii=False, separators=(',', ':'))

print(f"\n=== SUMMARY ===")
print(f"Added: {added}")
print(f"Skipped (duplicates): {skipped}")
print(f"Total records: {len(restaurants)}")
